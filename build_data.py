#!/usr/bin/env python3
"""
build_data.py - Builds platform data from:
  - Layer 1 (Policies): Mex obs .xlsx
  - Layers 2-6 (Federal, Investments, Green Mfg, Mines, Polos): Observatory RA Dataset - April 2026.xlsx
Run from project root: python3 build_data.py
"""

import openpyxl
import json
import os
import re
from datetime import datetime, date

MASTER = "../Mex obs .xlsx"
RA = "../Observatory RA Dataset - April 2026.xlsx"
GEOJSON = "data/mexico-states.geojson"
OUTPUT_DIR = "data/"
EMBEDDED_JS = "js/embedded-data.js"

INSTRUMENT_MAP = {
    1: "Grant", 2: "Loan", 3: "Tax relief", 4: "Subsidy/Support",
    5: "Loan guarantee", 6: "Joint venture", 7: "Public procurement",
    8: "Infrastructure", 9: "Coordination mechanism", 10: "FDI measures",
    11: "Workforce development"
}

STATE_NAME_MAP = {
    "Coahuila de Zaragoza": "Coahuila",
    "Distrito Federal": "Ciudad de México",
    "México": "Estado de México",
    "Mexico": "Estado de México",
    "Mexico City": "Ciudad de México",
    "Ciudad de Mexico": "Ciudad de México",
    "Estado de Mexico": "Estado de México",
    "Michoacán de Ocampo": "Michoacán",
    "Michoacan": "Michoacán",
    "Veracruz de Ignacio de la Llave": "Veracruz",
    "State of Mexico": "Estado de México",
    "Edomex": "Estado de México",
    "Nuevo Leon": "Nuevo León",
    "Queretaro": "Querétaro",
    "San Luis Potosi": "San Luis Potosí",
    "Yucatan": "Yucatán",
}

# Polo coordinates (from existing platform data)
POLO_COORDS = {
    "Coatzacoalcos I": ("Veracruz", 18.1341, -94.4588),
    "Coatzacoalcos II": ("Veracruz", 18.15, -94.43),
    "Texistepec": ("Veracruz", 17.9069, -94.8131),
    "San Juan Evangelista": ("Veracruz", 17.89, -95.14),
    "Matías Romero (Donají)": ("Oaxaca", 16.8772, -95.0395),
    "Ixtaltepec (Chivela)": ("Oaxaca", 16.5139, -95.062),
    "San Blas Atempa": ("Oaxaca", 16.3248, -95.2229),
    "Santa María Mixtequilla": ("Oaxaca", 16.45, -95.15),
    "Ciudad Ixtepec": ("Oaxaca", 16.5622, -95.1042),
    "Salina Cruz": ("Oaxaca", 16.1711, -95.1967),
    "Teapa": ("Tabasco", 17.5469, -92.9528),
    "Tapachula I y II": ("Chiapas", 14.9039, -92.2572),
    "Seybaplaya": ("Campeche", 19.6478, -90.6928),
    "Ciudad Juárez": ("Chihuahua", 31.6904, -106.4245),
    "Durango": ("Durango", 24.0277, -104.6532),
    "Nezahualcóyotl": ("Estado de México", 19.4003, -99.0145),
    "Celaya": ("Guanajuato", 20.5283, -100.8157),
    "Zapotlán (AIFA)": ("Hidalgo", 19.8428, -98.9206),
    "San José Chiapa (Ciudad Modelo)": ("Puebla", 19.3461, -97.7806),
    "Topolobampo": ("Sinaloa", 25.601, -109.05),
    "Altamira": ("Tamaulipas", 22.3933, -97.9431),
    "Huamantla": ("Tlaxcala", 19.3142, -97.9233),
    "Tuxpan": ("Veracruz", 20.9553, -97.3986),
    "Morelia-Zinapécuaro": ("Michoacán", 19.8617, -100.8333),
    "Chetumal": ("Quintana Roo", 18.5022, -88.2961),
    "Tula": ("Hidalgo", 20.0534, -99.3438),
    "Hermosillo": ("Sonora", 29.0729, -110.9559),
    "ZM Mérida": ("Yucatán", 20.9674, -89.5926),
    "Lázaro Cárdenas": ("Guerrero", 17.9581, -101.8583),
    "Zona Carbonífera": ("Coahuila", 27.85, -101.1167),
    "Puebla": ("Puebla", 19.34, -97.78),
}

# Official PODECOBI status updates (en.podecobi.com + Opportimes May 2025)
POLO_STATUS = {
    "Coatzacoalcos I": "Operational",
    "Coatzacoalcos II": "Operational",
    "Texistepec": "Operational",
    "San Juan Evangelista": "Operational",
    "Matías Romero (Donají)": "Operational",
    "Ixtaltepec (Chivela)": "Operational",
    "San Blas Atempa": "Operational",
    "Santa María Mixtequilla": "Operational",
    "Ciudad Ixtepec": "Operational",
    "Salina Cruz": "Operational",
    "Tapachula I y II": "Under construction",
    "Teapa": "Under construction",
    "Ciudad Juárez": "Operational",
    "Zapotlán (AIFA)": "Operational",
    "San José Chiapa (Ciudad Modelo)": "Operational",
    "Huamantla": "Operational",
    "Morelia-Zinapécuaro": "Operational",
    "ZM Mérida": "Approved/Planned",
    "Lázaro Cárdenas": "Approved/Planned",
    "Zona Carbonífera": "Approved/Planned",
    "Hermosillo": "Under construction",
}

POLO_CATEGORY = {
    "Operational": "En marcha",
    "Under construction": "En proceso",
    "Announced": "Nuevos polos",
    "Approved/Planned": "En evaluación",
}


def normalize_state(name):
    if not name:
        return None
    name = str(name).strip()
    if name in STATE_NAME_MAP:
        return STATE_NAME_MAP[name]
    if name in ('Multiple', 'National', 'Various', 'Unknown', 'N/A'):
        return None
    if ';' in name:
        first = name.split(';')[0].strip()
        return normalize_state(first)
    return name


def extract_year(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.year
    if isinstance(val, (int, float)):
        v = int(val)
        if 1990 <= v <= 2030:
            return v
        return None
    s = str(val).strip()
    m = re.match(r'(\d{4})', s)
    if m:
        y = int(m.group(1))
        if 1990 <= y <= 2030:
            return y
    return None


def serialize_value(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.isoformat()[:10]
    if isinstance(val, bool):
        return val
    if isinstance(val, float):
        if val != val:
            return None
        if val == int(val):
            return int(val)
        return round(val, 6)
    return val


def safe_int(val, default=0):
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def safe_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


# ============================================================
# LAYER 1: POLICIES (from Mex obs .xlsx)
# ============================================================
def build_policies(wb):
    print("Layer 1: Policies (from Mex obs .xlsx)...")
    ws = wb['Subnational ind policies']
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    headers = [str(h).strip().lower() if h else f'col_{i}' for i, h in enumerate(rows[0])]
    output = []
    for row in rows[1:]:
        if not row[0]:
            continue
        r = dict(zip(headers, [serialize_value(v) for v in row]))
        rec = {
            'id': int(r['id']),
            'state': normalize_state(r.get('state')),
            'name': str(r.get('name', '')).strip(),
            'description': r.get('description', ''),
            'instrument': safe_int(r.get('instrument'), None),
            'sector': r.get('sector'),
            'sector_id': str(r.get('sector_id', '')) if r.get('sector_id') else None,
            'hs_code': str(r.get('hs_code', '')) if r.get('hs_code') else None,
            'inv_govt': r.get('inv_govt'),
            'inv_private': r.get('inv_private'),
            'inv_currency': r.get('inv_currency'),
            'jobs': r.get('jobs'),
            'planmex': safe_int(r.get('planmex')),
            'polo': r.get('polo'),
            'announcement_date': r.get('announcement_date'),
            'implementation_start': r.get('implementation_start'),
            'implementation_end': r.get('implementation_end'),
            'green': safe_int(r.get('green')),
            'has_training': safe_int(r.get('has_training')),
            'r&d': safe_int(r.get('r&d')),
            'legal_basis': r.get('legal_basis'),
            'agency': r.get('agency'),
            'partners': r.get('partners'),
            'notes': r.get('notes'),
            'sources': r.get('sources'),
            'conditionality': safe_int(r.get('conditionality')),
            'pdp_typology': r.get('pdp_typology'),
            'inv_govt_millions': r.get('inv_govt_millions'),
            'inv_private_millions': r.get('inv_private_millions'),
            'status': r.get('status', 'Operational'),
            'hs_section': r.get('hs_section'),
            'hs_section_name': r.get('hs_section_name'),
            'pdp_typology_code': r.get('pdp_typology_code'),
            '_year': extract_year(r.get('announcement_date')),
        }
        output.append(rec)
    print(f"  -> {len(output)} policies")
    return output


# ============================================================
# LAYER 2: FEDERAL PROGRAMS (from RA Dataset)
# ============================================================
def build_federal(wb):
    print("Layer 2: Federal Programs (from RA Dataset)...")
    ws = wb["2. Federal-State Projects"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    headers_raw = rows[0]
    output = []
    idx = 1
    for row in rows[1:]:
        if not any(row):
            continue
        # Columns: coder(0), id(1), name(2), implementing_agency(3), description_short(4),
        # geographic_scope(5), states(6), sector(7), sector_id(8), instrument(9),
        # budget_investment(10), description(11), planmex(12), status(13), year_created(14),
        # sources(15), notes(16)
        name = row[2]
        if not name:
            continue
        year_created = safe_int(row[14], None)
        instrument = safe_int(row[9], None)
        rec = {
            'id': idx,
            'name': str(name).strip(),
            'implementing_agency': serialize_value(row[3]),
            'program_type': serialize_value(row[4]),
            'geographic_scope': serialize_value(row[5]),
            'states': str(row[6]) if row[6] else '',
            'sector': serialize_value(row[7]),
            'instrument': instrument,
            'budget_investment': serialize_value(row[10]),
            'description': serialize_value(row[11]),
            'planmex': safe_int(row[12]),
            'sources': serialize_value(row[15]),
            'instrument_text': INSTRUMENT_MAP.get(instrument, '') if instrument else '',
            'sector_original': serialize_value(row[7]),
            'sector_id': str(row[8]) if row[8] else None,
            'status': serialize_value(row[13]) or 'Operational',
            'year_created': year_created,
            'period_entry': year_created if year_created and year_created >= 2015 else 2015,
            '_year': year_created,
        }
        output.append(rec)
        idx += 1
    print(f"  -> {len(output)} federal programs")
    return output


# ============================================================
# LAYER 3: INVESTMENTS (from RA Dataset)
# ============================================================
def build_investments(wb, policies):
    print("Layer 3: Investments (from RA Dataset)...")
    ws = wb["3. Subnational Investments"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    output = []
    idx = 1
    for row in rows[1:]:
        if not any(row):
            continue
        # Columns: coder(0), id(1), company_name(2), origin_type(3), country_of_origin(4),
        # state(5), municipality_city(6), sector(7), sector_id(8), investment_amount(9),
        # investment_currency(10), jobs(11), year_announced(12), green(13), green_tech_id(14),
        # green_tech(15), green_reason(16), green_mfg_id(17), planmex(18),
        # description(19), status(20), sources(21), notes(22)
        name = row[2]
        if not name:
            continue
        year_announced = safe_int(row[12], None)
        rec = {
            'id': idx,
            'company_name': str(name).strip(),
            'country_of_origin': serialize_value(row[4]),
            'state': normalize_state(row[5]),
            'municipality_city': serialize_value(row[6]),
            'sector': serialize_value(row[7]),
            'sector_id': str(row[8]) if row[8] else None,
            'investment_usd_millions': safe_float(row[9]),
            'jobs': safe_int(row[11], None),
            'year_announced': year_announced,
            'green': safe_int(row[13]),
            'green_tech_id': safe_int(row[14], None),
            'green_tech': serialize_value(row[15]),
            'green_reason': serialize_value(row[16]),
            'planmex': safe_int(row[18]),
            'nearshoring': 0,
            'description': serialize_value(row[19]),
            'sources': serialize_value(row[21]),
            'status': serialize_value(row[20]) or 'Operational',
            '_year': year_announced,
            'linked_policy_ids': [],
            'linkage_confidence': [],
            'linkage_basis': [],
        }
        output.append(rec)
        idx += 1

    # Compute linkages
    print("  Computing investment-policy linkages...")
    for inv in output:
        inv_state = inv.get('state')
        inv_sector = str(inv.get('sector_id', '')) if inv.get('sector_id') else ''
        if not inv_state or not inv_sector:
            continue
        for pol in policies:
            pol_state = pol.get('state')
            pol_sector = str(pol.get('sector_id', '')) if pol.get('sector_id') else ''
            if not pol_state or not pol_sector:
                continue
            if pol_state == inv_state and inv_sector[:4] == pol_sector[:4]:
                inv['linked_policy_ids'].append(pol['id'])
                if inv_sector == pol_sector:
                    inv['linkage_confidence'].append('high')
                    inv['linkage_basis'].append(f"Exact SCIAN: {pol_sector}")
                else:
                    inv['linkage_confidence'].append('medium')
                    inv['linkage_basis'].append(f"SCIAN prefix: {pol_sector}")

    linked = sum(1 for f in output if f['linked_policy_ids'])
    print(f"  -> {len(output)} investments ({linked} linked to policies)")
    return output


# ============================================================
# LAYER 4: GREEN MANUFACTURING (from RA Dataset)
# ============================================================
def build_greenmfg(wb):
    print("Layer 4: Green Manufacturing (from RA Dataset)...")
    ws = wb["4. Green Manufacturing"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    output = []
    idx = 1
    for row in rows[1:]:
        if not any(row):
            continue
        # Columns: coder(0), id(1), name(2), company(3), foreign(4), country_of_origin(5),
        # sector(6), product(7), city(8), state(9), state_code(10), latitude(11),
        # longitude(12), investment_amount(13), investment_currency(14),
        # announcement_date(15), production_start_date(16), target_year(17),
        # target_production(18), production_units(19), post_ira(20),
        # sector_id(21), hs_code(22), status(23), sources(24)
        name = row[2]
        if not name:
            continue
        announcement_date = serialize_value(row[15])
        target_year = safe_int(row[17], None)
        rec = {
            'id': idx,
            'name': str(name).strip(),
            'company': serialize_value(row[3]),
            'sector': serialize_value(row[6]),
            'product': serialize_value(row[7]),
            'city': serialize_value(row[8]),
            'state': normalize_state(row[9]),
            'state_code': serialize_value(row[10]),
            'latitude': safe_float(row[11]),
            'longitude': safe_float(row[12]),
            'investment_usd_millions': safe_float(row[13]),
            'announcement_date': announcement_date,
            'production_start_date': serialize_value(row[16]),
            'target_year': target_year,
            'target_production': serialize_value(row[18]),
            'realized_production': None,
            'production_units': serialize_value(row[19]),
            'post_ira': safe_int(row[20]),
            'sources': serialize_value(row[24]) if len(row) > 24 else None,
            'status': serialize_value(row[23]) if len(row) > 23 else 'Operational',
            '_year': extract_year(announcement_date) or target_year,
        }
        output.append(rec)
        idx += 1
    print(f"  -> {len(output)} green manufacturing records")
    return output


# ============================================================
# LAYER 5: MINES (from RA Dataset)
# ============================================================
def build_mines(wb):
    print("Layer 5: Mines (from RA Dataset)...")
    ws = wb["5. Mines and Smelters"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    output = []
    idx = 1
    for row in rows[1:]:
        if not any(row):
            continue
        # Columns: coder(0), id(1), property_name(2), property_id(3), primary_commodity(4),
        # commodity(5), facility_type(6), development_stage(7), capital_cost(8),
        # sustaining_cost(9), study_type(10), study_year(11), latitude(12), longitude(13),
        # state(14), municipality(15), geocode_method(16), status(17), sources(18), notes(19)
        property_name = row[2]
        if not property_name:
            continue
        rec = {
            'id': idx,
            'property_name': str(property_name).strip(),
            'property_id': serialize_value(row[3]),
            'primary_commodity': serialize_value(row[4]),
            'commodity': serialize_value(row[5]),
            'facility_type': serialize_value(row[6]),
            'development_stage': serialize_value(row[7]),
            'capital_cost_usd_millions': safe_float(row[8]),
            'sustaining_cost_usd_millions': safe_float(row[9]),
            'study_type': serialize_value(row[10]),
            'study_year': safe_int(row[11], None),
            'latitude': safe_float(row[12]),
            'longitude': safe_float(row[13]),
            'state': normalize_state(row[14]),
            'municipality': serialize_value(row[15]),
            'geocode_method': serialize_value(row[16]),
            'status': serialize_value(row[17]) or 'Operational',
            'sources': serialize_value(row[18]),
            '_year': safe_int(row[11], None),
        }
        output.append(rec)
        idx += 1
    print(f"  -> {len(output)} mines")
    return output


# ============================================================
# LAYER 6: POLOS (from RA Dataset + geocoding + official updates)
# ============================================================
def build_polos(wb):
    print("Layer 6: Polos (from RA Dataset + official PODECOBI data)...")
    ws = wb["6. Polos de Bienestar"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    output = []
    idx = 1
    for row in rows[1:]:
        if not any(row):
            continue
        # Columns: coder(0), id(1), name(2), description(3), municipality(4), state(5),
        # sector_id(6), hs_code(7), inv_govt(8), inv_private(9), inv_currency(10),
        # jobs(11), area_hectares(12), operator(13), operational_date(14),
        # fiscal_end_date(15), zee_predecessor(16), planmex(17), green(18),
        # status(19), legal_basis(20), sources(21), notes(22)
        name = row[2]
        if not name:
            continue
        name = str(name).strip()

        # Get coordinates and state from lookup
        coords = POLO_COORDS.get(name)
        if coords:
            state, lat, lon = coords
        else:
            state = normalize_state(row[5]) or 'Unknown'
            lat, lon = None, None

        # Get official status if available
        status = POLO_STATUS.get(name, serialize_value(row[19]) or 'Announced')
        category = POLO_CATEGORY.get(status, 'Nuevos polos')

        # Parse sectors from sector_id field
        sector_ids = []
        if row[6]:
            sector_ids = [s.strip() for s in str(row[6]).split(';') if s.strip()]

        rec = {
            'id': idx,
            'name': name,
            'name_official': f"Polo de Desarrollo para el Bienestar {name}",
            'category': category,
            'state': state,
            'municipality': serialize_value(row[4]),
            'latitude': lat,
            'longitude': lon,
            'status': status,
            'decree_date': '2025-05-22',
            'announcement_date': serialize_value(row[14]) or '2024-06-01',
            'sectors': sector_ids,
            'sector_ids': sector_ids,
            'infrastructure_type': None,
            'ciit_connected': idx <= 10,
            'planmex': safe_int(row[17], 1),
            'investment_committed_usd_millions': safe_float(row[8]),
            'investment_projected_usd_millions': safe_float(row[9]),
            'investment_source': None,
            'jobs_projected': safe_int(row[11], None),
            'companies_announced': [],
            'area_hectares': safe_float(row[12]),
            'tax_incentive_active': True,
            'incentive_expiry_year': extract_year(row[15]),
            'linked_policy_ids': [],
            'linked_fdi_ids': [],
            'former_zee': bool(safe_int(row[16])),
            'former_zee_name': None,
            'phase': None,
            'legal_basis': serialize_value(row[20]),
            'sources': serialize_value(row[21]) or 'Gobierno de Mexico. 2025. "Decreto por el que se establecen los Polos de Desarrollo Economico para el Bienestar." Diario Oficial de la Federacion, May 22. Mexico City: Secretaria de Gobernacion. https://en.podecobi.com',
            'notes': serialize_value(row[22]) if len(row) > 22 else None,
            '_year': 2024,
        }
        output.append(rec)
        idx += 1
    print(f"  -> {len(output)} polos")
    return output


# ============================================================
# OUTPUT
# ============================================================
def write_json(data, filename):
    path = os.path.join(OUTPUT_DIR, filename)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, separators=(',', ':'))
    size_kb = os.path.getsize(path) / 1024
    print(f"  Written: {path} ({size_kb:.0f} KB)")


def write_embedded_js(policies, federal, investments, greenmfg, mines, polos):
    print("Generating embedded-data.js...")
    with open(GEOJSON, 'r', encoding='utf-8') as f:
        geojson = json.load(f)

    timestamp = datetime.now().isoformat()
    parts = [
        f"// Auto-generated embedded data - do not edit manually",
        f"// Generated: {timestamp}",
        f"// Layer 1: Mex obs .xlsx | Layers 2-6: Observatory RA Dataset - April 2026.xlsx",
        "",
        f"const RAW_POLICIES = {json.dumps(policies, ensure_ascii=False)};",
        "",
        f"const RAW_FEDERAL = {json.dumps(federal, ensure_ascii=False)};",
        "",
        f"const RAW_FDI = {json.dumps(investments, ensure_ascii=False)};",
        "",
        f"const RAW_GREENMFG = {json.dumps(greenmfg, ensure_ascii=False)};",
        "",
        f"const RAW_MINES = {json.dumps(mines, ensure_ascii=False)};",
        "",
        f"const RAW_POLOS = {json.dumps(polos, ensure_ascii=False)};",
        "",
        f"const RAW_GEOJSON = {json.dumps(geojson, ensure_ascii=False)};",
        "",
    ]

    with open(EMBEDDED_JS, 'w', encoding='utf-8') as f:
        f.write('\n'.join(parts))

    size_kb = os.path.getsize(EMBEDDED_JS) / 1024
    print(f"  Written: {EMBEDDED_JS} ({size_kb:.0f} KB)")


def main():
    print("=" * 60)
    print("Mexico Observatory Data Build")
    print("=" * 60)

    print(f"\nLayer 1 source: {MASTER}")
    master_wb = openpyxl.load_workbook(MASTER, read_only=True, data_only=True)

    print(f"Layers 2-6 source: {RA}")
    ra_wb = openpyxl.load_workbook(RA, read_only=True, data_only=True)

    print()
    policies = build_policies(master_wb)
    federal = build_federal(ra_wb)
    investments = build_investments(ra_wb, policies)
    greenmfg = build_greenmfg(ra_wb)
    mines = build_mines(ra_wb)
    polos = build_polos(ra_wb)

    master_wb.close()
    ra_wb.close()

    print("\nWriting JSON files...")
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    write_json(policies, 'policies.json')
    write_json(federal, 'federal.json')
    write_json(investments, 'fdi.json')
    write_json(greenmfg, 'greenmfg.json')
    write_json(mines, 'mines.json')
    write_json(polos, 'polos.json')

    print()
    write_embedded_js(policies, federal, investments, greenmfg, mines, polos)

    print("\n" + "=" * 60)
    print("BUILD COMPLETE")
    print(f"  Layer 1 - Policies:       {len(policies)}")
    print(f"  Layer 2 - Federal:        {len(federal)}")
    print(f"  Layer 3 - Investments:    {len(investments)}")
    print(f"  Layer 4 - Green Mfg:      {len(greenmfg)}")
    print(f"  Layer 5 - Mines:          {len(mines)}")
    print(f"  Layer 6 - Polos:          {len(polos)}")
    print(f"  Energy:                   DROPPED")
    print("=" * 60)


if __name__ == '__main__':
    main()
